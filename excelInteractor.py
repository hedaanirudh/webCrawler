from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table


class Excel:

    def __write_to_sheet(self, ws, df, row_no=1, col_no=1):
        rows = dataframe_to_rows(df, index=False)
        for r_idx, row in enumerate(rows, row_no):
            for c_idx, value in enumerate(row, col_no):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Create excel workbook with data sheet and index words sheet
    def create_workbook(self, final_stats, data_df, index_dict, time_tracker_dict):
        column_widths = [15, 70, 150, 15]
        wb = Workbook()
        # Sheet for Crawler Data
        ws = wb.active
        ws.title = "Crawler Data"
        self.__write_to_sheet(ws, data_df)
        for i in range(1, 5):
            ws.column_dimensions[get_column_letter(i)].width = column_widths[i - 1]
        table = Table(displayName="Table1", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))
        ws.add_table(table)

        # Sheet for Indexed Words
        ws1 = wb.create_sheet("Index Words")
        ws1["A1"] = "Index Word"
        ws1["B1"] = "Document Ids"
        for row, (index, ids) in enumerate(index_dict.items(), start=2):
            # ws1[f"A{row}"] = index.encode('unicode_escape').decode('utf-8')
            ws1[f"A{row}"] = index
            ws1[f"B{row}"] = ','.join(ids)
        for i in range(1, 3):
            ws1.column_dimensions[get_column_letter(i)].width = column_widths[i - 1]
        table = Table(displayName="Table2", ref="A1:" + get_column_letter(ws1.max_column) + str(ws1.max_row))
        ws1.add_table(table)

        # Stats per 50 docs
        row = 25
        ws3 = wb.create_sheet("Stats")
        ws3["A1"] = "# Pages Crawled"
        ws3["B1"] = "Time For last 100(s)"
        ws3["C1"] = "Total Time(s)"
        ws3["D1"] = "Total Keywords Encountered"
        ws3["E1"] = "Average Keywords Per Page"
        ws3["F1"] = "Total Unique Keywords Encountered"
        ws3["G1"] = "Average Unique Keywords Per Page"
        ws3["H1"] = "URLs To Be Crawled"
        for row, (docs, value) in enumerate(time_tracker_dict.items(), start=2):
            ws3[f"A{row}"] = docs
            ws3[f"B{row}"] = value["last_time"]
            ws3[f"C{row}"] = value["total_time"]
            ws3[f"D{row}"] = value["total_keywords"]
            ws3[f"E{row}"] = value["average_keywords"]
            ws3[f"F{row}"] = value["total_unique_keywords"]
            ws3[f"G{row}"] = value["average_unique_keywords"]
            ws3[f"H{row}"] = value["urls_to_be_crawled"]

        for i in range(1, 9):
            ws3.column_dimensions[get_column_letter(i)].width = 30

        # Stats Summary
        for r, (key, value) in enumerate(final_stats.items(), start=row + 5):
            ws3[f"A{r}"] = key
            ws3[f"B{r}"] = value

        wb.save('web_crawler_output.xlsx')